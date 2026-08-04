import json

from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.contrib import messages
from django.contrib.sessions.models import Session
from django.db.models import Q
from .models import DetailVente, Vente
from apps.menu.models import Produit
from apps.stock.models import MouvementStock, Stock
from apps.accounts.decorators import role_required
from apps.accounts.models import CustomUser
from apps.parametres.models import ParametreRestaurant



@role_required(["CAISSIER"])
def pos(request):
    from apps.menu.models import Categorie
    produits = Produit.objects.filter(disponible=True).select_related('categorie', 'stock')
    categories = Categorie.objects.all()
    parametre = ParametreRestaurant.load()

    return render(
        request,
        "ventes/pos.html",
        {
            "produits": produits,
            "categories": categories,
            "parametre": parametre,
        }
    )


@role_required(["CAISSIER"])
def enregistrer_vente(request):
    if request.method != "POST":
        return JsonResponse({
            "success": False,
            "message": "Méthode non autorisée."
        }, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({
            "success": False,
            "message": "Données invalides."
        }, status=400)

    panier = data.get("panier", [])
    mode = data.get("mode_paiement")

    if not panier:
        return JsonResponse({
            "success": False,
            "message": "Le panier est vide."
        }, status=400)

    ids_produits = [int(item.get("id")) for item in panier]
    produits = {
        produit.id: produit
        for produit in Produit.objects.filter(id__in=ids_produits)
    }

    for item in panier:
        produit_id = int(item.get("id"))
        produit = produits.get(produit_id)

        if produit is None:
            return JsonResponse({
                "success": False,
                "message": "Produit introuvable."
            }, status=404)

        quantite = int(item.get("qte", 0))

        if quantite <= 0:
            return JsonResponse({
                "success": False,
                "message": f"Quantité invalide : {produit.nom}"
            }, status=400)

        stock = Stock.objects.filter(produit=produit).first()

        if stock is None or stock.quantite < quantite:
            return JsonResponse({
                "success": False,
                "message": f"Stock insuffisant : {produit.nom}"
            }, status=400)

    total = 0

    with transaction.atomic():
        vente = Vente.objects.create(
            caissier=request.user,
            total=0,
            mode_paiement=mode
        )

        for item in panier:
            produit = produits[int(item["id"])]
            quantite = int(item["qte"])
            prix = produit.prix
            sous_total = prix * quantite
            total += sous_total

            DetailVente.objects.create(
                vente=vente,
                produit=produit,
                quantite=quantite,
                prix=prix,
                sous_total=sous_total
            )

            stock = Stock.objects.select_for_update().filter(produit=produit).first()


            if not stock:
                raise Exception(f"Stock absent pour {produit.nom}" )
            stock.quantite -= quantite
            stock.save()

            MouvementStock.objects.create(
                produit=produit,
                type_mouvement="SORTIE",
                quantite=quantite,
                utilisateur=request.user,
                commentaire=f"Vente N° {vente.id}"
            )

        vente.total = total
        vente.save()

    return JsonResponse({
        "success": True,
        "vente_id": vente.id
    })


@role_required(["ADMIN", "CAISSIER"])
def ticket(request, vente_id):
    vente = get_object_or_404(
        Vente,
        id=vente_id
    )
    parametre = ParametreRestaurant.load()

    return render(
        request,
        "ventes/ticket.html",
        {
            "vente": vente,
            "parametre": parametre,
        }
    )


@role_required(["ADMIN", "CAISSIER"])
def detail_vente(request, vente_id):
    vente = get_object_or_404(
        Vente.objects.select_related("caissier", "annule_par"),
        id=vente_id
    )
    parametre = ParametreRestaurant.load()

    return render(
        request,
        "ventes/detail.html",
        {
            "vente": vente,
            "parametre": parametre,
        }
    )


@role_required(["CAISSIER"])
def annuler_vente(request, vente_id):
    """Annule une vente et remet les produits en stock."""
    vente = get_object_or_404(Vente, id=vente_id)

    if vente.annulee:
        messages.error(request, "Cette vente a déjà été annulée.")
        return redirect("ventes:detail", vente_id=vente.id)

    if request.method == "POST":
        with transaction.atomic():
            vente.annulee = True
            vente.annule_le = timezone.now()
            vente.annule_par = request.user
            vente.save()

            for detail in vente.details.all():
                stock, _ = Stock.objects.get_or_create(produit=detail.produit)
                stock.quantite += detail.quantite
                stock.save()

                MouvementStock.objects.create(
                    produit=detail.produit,
                    type_mouvement="ENTREE",
                    quantite=detail.quantite,
                    utilisateur=request.user,
                    commentaire=f"Retour stock — annulation vente N° {vente.id}",
                )

        messages.success(
            request,
            f"Vente N° {vente.id} annulée. Les produits ont été remis en stock."
        )
        return redirect("ventes:historique")

    return render(request, "ventes/annuler.html", {"vente": vente})


@role_required(["ADMIN", "CAISSIER"])
def historique(request):

    ventes = _ventes_filtrees(request)

    sessions = Session.objects.filter(
        expire_date__gte=timezone.now()
    )

    users_online = []

    for session in sessions:
        data = session.get_decoded()
        user_id = data.get("_auth_user_id")
        if user_id:
            users_online.append(user_id)

    users_online = list(set(users_online))

    utilisateurs_connectes = CustomUser.objects.filter(
        id__in=users_online,
        is_active=True
    )

    total = sum(
        vente.total for vente in ventes
    )

    return render(
        request,
        "ventes/historique.html",
        {
            "ventes": ventes,
            "total": total,
            "utilisateurs_connectes": utilisateurs_connectes,
            "nombre_connectes": utilisateurs_connectes.count(),
            "q": request.GET.get("q", "").strip(),
            "date": request.GET.get("date", "").strip(),
            "statut": request.GET.get("statut", "").strip(),
        }
    )


def _ventes_filtrees(request):
    """Applique les mêmes filtres que l'historique (q, date, statut)."""
    ventes = Vente.objects.select_related("caissier").order_by("-created_at")

    q = request.GET.get("q", "").strip()
    date = request.GET.get("date", "").strip()
    statut = request.GET.get("statut", "").strip()

    if q:
        ventes = ventes.filter(
            Q(id__icontains=q) | Q(caissier__username__icontains=q)
        )
    if date:
        ventes = ventes.filter(created_at__date=date)
    if statut == "annulees":
        ventes = ventes.filter(annulee=True)
    elif statut == "valides":
        ventes = ventes.filter(annulee=False)

    return ventes


@role_required(["ADMIN", "CAISSIER"])
def exporter_excel(request):
    """Exporte l'historique des ventes en fichier Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    ventes = _ventes_filtrees(request)
    parametre = ParametreRestaurant.load()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes"

    entete_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    entete_font = Font(color="FFFFFF", bold=True)

    # Titre
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{parametre.nom} - Historique des ventes"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Exporté le {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    entetes = ["N° Ticket", "Caissier", "Date", "Paiement", "Total (FCFA)", "Statut"]
    for col, titre in enumerate(entetes, start=1):
        cell = ws.cell(row=4, column=col, value=titre)
        cell.fill = entete_fill
        cell.font = entete_font
        cell.alignment = Alignment(horizontal="center")

    for i, vente in enumerate(ventes, start=5):
        ws.cell(row=i, column=1, value=vente.id)
        ws.cell(row=i, column=2, value=vente.caissier.username if vente.caissier else "—")
        ws.cell(row=i, column=3, value=timezone.localtime(vente.created_at).strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=i, column=4, value=vente.get_mode_paiement_display())
        ws.cell(row=i, column=5, value=vente.total)
        ws.cell(row=i, column=6, value="Annulée" if vente.annulee else "Valide")

    # Largeurs de colonnes
    for col, largeur in zip("ABCDEF", [14, 18, 18, 14, 16, 12]):
        ws.column_dimensions[col].width = largeur

    # Total
    total_ligne = len(ventes) + 5
    ws.cell(row=total_ligne, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_ligne, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=total_ligne, column=5, value=sum(v.total for v in ventes)).font = Font(bold=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="ventes_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    wb.save(response)
    return response


@role_required(["ADMIN", "CAISSIER"])
def exporter_pdf(request):
    """Exporte l'historique des ventes en fichier PDF."""
    from fpdf import FPDF

    ventes = _ventes_filtrees(request)
    parametre = ParametreRestaurant.load()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # En-tête
    pdf.set_font("helvetica", "B", 16)
    pdf.set_text_color(249, 115, 22)
    pdf.cell(0, 10, parametre.nom, ln=True, align="C")
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(0, 8, "Historique des ventes", ln=True, align="C")
    pdf.set_font("helvetica", "", 9)
    pdf.set_text_color(100, 116, 139)
    pdf.cell(0, 6, f"Exporté le {timezone.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align="C")
    pdf.ln(5)

    # Tableau
    colonnes = ("N°", "Caissier", "Date", "Paiement", "Total", "Statut")
    largeurs = (14, 34, 34, 30, 30, 30)
    pdf.set_font("helvetica", "B", 9)
    pdf.set_fill_color(249, 115, 22)
    pdf.set_text_color(255, 255, 255)
    for col, largeur in zip(colonnes, largeurs):
        pdf.cell(largeur, 8, col, border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(30, 41, 59)
    pdf.set_font("helvetica", "", 8)
    total_ventes = 0
    for vente in ventes:
        ligne = (
            str(vente.id),
            vente.caissier.username if vente.caissier else "—",
            timezone.localtime(vente.created_at).strftime("%d/%m/%Y %H:%M"),
            vente.get_mode_paiement_display(),
            f"{vente.total:,}".replace(",", " "),
            "Annulée" if vente.annulee else "Valide",
        )
        x_depart = pdf.get_x()
        if pdf.get_y() > 250:
            pdf.add_page()
            x_depart = 10
        for texte, largeur in zip(ligne, largeurs):
            pdf.cell(largeur, 7, texte, border=1, align="C")
        pdf.ln()
        total_ventes += vente.total

    # Total
    pdf.set_font("helvetica", "B", 9)
    pdf.set_text_color(249, 115, 22)
    pdf.cell(4 * sum(largeurs[:4]) / 4, 0, "")
    total_largeur = sum(largeurs[:5])
    pdf.set_font("helvetica", "B", 9)
    pdf.set_text_color(30, 41, 59)
    pdf.set_fill_color(255, 237, 213)
    pdf.cell(total_largeur, 8, f"TOTAL : {total_ventes:,} FCFA".replace(",", " "), border=1, fill=True, align="R")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ventes_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    pdf.output(response)
    return response